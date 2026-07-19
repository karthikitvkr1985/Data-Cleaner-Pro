"""Structure detection — find tables, headers, and merged cells in Excel files."""
from __future__ import annotations

import io
from dataclasses import dataclass

from backend.models.schemas import TableDetectionResult, TableInfo


def detect_tables(
    file_bytes: bytes,
    file_ext: str,
    sheet_name: str,
    header_row: int | None = None,
    start_row: int | None = None,
) -> TableDetectionResult:
    """Detect table boundaries and header rows in an Excel or CSV file."""
    if file_ext in (".csv", ".tsv"):
        return _detect_csv_structure(file_bytes, file_ext)
    return _detect_excel_structure(file_bytes, sheet_name, header_row, start_row)


def _detect_csv_structure(file_bytes: bytes, ext: str) -> TableDetectionResult:
    import pandas as pd

    sep = "\t" if ext == ".tsv" else ","
    try:
        df = pd.read_csv(io.BytesIO(file_bytes), sep=sep, nrows=5)
        nrows = len(pd.read_csv(io.BytesIO(file_bytes), sep=sep))
        ncols = len(df.columns)
    except Exception:
        return TableDetectionResult(sheet_name="Sheet1", tables=[])

    table = TableInfo(
        start_row=0,
        end_row=nrows - 1,
        start_col=0,
        end_col=ncols - 1,
        header_row=0,
        confidence=0.95,
    )
    return TableDetectionResult(sheet_name="Sheet1", tables=[table])


def _detect_excel_structure(
    file_bytes: bytes,
    sheet_name: str,
    forced_header: int | None,
    forced_start: int | None,
) -> TableDetectionResult:
    try:
        import openpyxl
    except ImportError:
        return TableDetectionResult(sheet_name=sheet_name, tables=[])

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=False, data_only=True)
    except Exception:
        return TableDetectionResult(sheet_name=sheet_name, tables=[])

    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    # Read raw cells into a 2D list
    rows: list[list] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))

    if not rows:
        wb.close()
        return TableDetectionResult(sheet_name=sheet_name, tables=[])

    # Propagate merged cell values
    merged_values: dict[tuple[int, int], object] = {}
    for merge in ws.merged_cells.ranges:
        top_left = ws.cell(merge.min_row, merge.min_col).value
        for r in range(merge.min_row, merge.max_row + 1):
            for c in range(merge.min_col, merge.max_col + 1):
                merged_values[(r - 1, c - 1)] = top_left
    for (r, c), val in merged_values.items():
        if r < len(rows) and c < len(rows[r]):
            rows[r][c] = val

    wb.close()

    # Find table boundaries using blank row/col separators
    tables = _find_table_boundaries(rows, forced_header, forced_start)
    return TableDetectionResult(sheet_name=sheet_name, tables=tables)


def _row_is_blank(row: list) -> bool:
    return all(v is None or str(v).strip() == "" for v in row)


def _row_text_density(row: list) -> float:
    non_null = sum(1 for v in row if v is not None and str(v).strip() != "")
    return non_null / len(row) if row else 0.0


def _row_is_likely_header(row: list, next_row: list | None) -> tuple[bool, float]:
    """Heuristic: header rows have high text density and row below has a type shift."""
    if not row:
        return False, 0.0
    density = _row_text_density(row)
    if density < 0.5:
        return False, 0.0

    # All string values → likely header
    string_count = sum(1 for v in row if v is not None and isinstance(v, str) and str(v).strip())
    string_ratio = string_count / len(row)

    confidence = string_ratio * 0.7 + density * 0.3

    if next_row is not None:
        # Check type shift: row is strings, next row has numeric values
        next_numeric = sum(1 for v in next_row if v is not None and isinstance(v, (int, float)))
        if next_numeric > 0:
            confidence = min(1.0, confidence + 0.2)

    return string_ratio > 0.6, confidence


def _find_col_bounds(rows: list[list], row_start: int, row_end: int) -> tuple[int, int]:
    max_col = max((len(r) for r in rows[row_start:row_end + 1]), default=0)
    # Find first non-blank col
    first_col = 0
    last_col = max_col - 1
    for ci in range(max_col):
        if any(
            ci < len(rows[ri]) and rows[ri][ci] is not None and str(rows[ri][ci]).strip()
            for ri in range(row_start, row_end + 1)
        ):
            first_col = ci
            break
    for ci in range(max_col - 1, -1, -1):
        if any(
            ci < len(rows[ri]) and rows[ri][ci] is not None and str(rows[ri][ci]).strip()
            for ri in range(row_start, row_end + 1)
        ):
            last_col = ci
            break
    return first_col, last_col


def _find_table_boundaries(
    rows: list[list],
    forced_header: int | None,
    forced_start: int | None,
) -> list[TableInfo]:
    n = len(rows)
    if n == 0:
        return []

    # Split by blank rows
    regions: list[tuple[int, int]] = []
    start = None
    consecutive_blank = 0
    BLANK_THRESHOLD = 2

    for i, row in enumerate(rows):
        if _row_is_blank(row):
            consecutive_blank += 1
            if consecutive_blank >= BLANK_THRESHOLD and start is not None:
                regions.append((start, i - consecutive_blank))
                start = None
        else:
            consecutive_blank = 0
            if start is None:
                start = i
    if start is not None:
        regions.append((start, n - 1))

    tables = []
    for r_start, r_end in regions:
        if r_end < r_start:
            continue

        # Determine header row
        header_row = forced_header
        confidence = 0.85

        if header_row is None:
            # Try each row in region as potential header
            best_conf = 0.0
            for ri in range(r_start, min(r_start + 5, r_end + 1)):
                next_r = rows[ri + 1] if ri + 1 <= r_end else None
                is_header, conf = _row_is_likely_header(rows[ri], next_r)
                if is_header and conf > best_conf:
                    best_conf = conf
                    header_row = ri
            confidence = best_conf if best_conf > 0 else 0.65

        first_col, last_col = _find_col_bounds(rows, r_start, r_end)
        tables.append(TableInfo(
            start_row=r_start,
            end_row=r_end,
            start_col=first_col,
            end_col=last_col,
            header_row=header_row,
            confidence=round(confidence, 3),
        ))

    return tables
